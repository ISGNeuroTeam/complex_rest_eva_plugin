from rest.permissions import IsAuthenticated
from rest.response import Response, status
from rest.views import APIView
from openpyxl import Workbook
from datetime import datetime
import super_logger
import tempfile
import tarfile
import uuid
import json
import os

from ..settings import STATIC_CONF, DB_CONN


class QuizExportJsonHandlerView(APIView):
    """
    There is method for export one or more quiz object in '.json' format files.
    Json files returns in 'tar.gz' package with uuid-name.
    """

    permission_classes = (IsAuthenticated,)
    # permission_classes = (AllowAny,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def _save_quiz2storage(self, quiz_ids):

        with tempfile.TemporaryDirectory() as tmp_dir:
            archive_name = f"{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.eva.quiz"
            _dirname = self.handler_id
            _base_path = os.path.join(STATIC_CONF['static_path'], 'storage', _dirname)
            if not os.path.exists(_base_path):
                os.makedirs(_base_path)

            archive_path = os.path.join(_base_path, archive_name)
            archive = tarfile.open(archive_path, mode='x:gz')

            for qid in quiz_ids:
                try:
                    quiz_data = DB_CONN.get_quiz(quiz_id=qid)
                    if not quiz_data:
                        return Response(
                            json.dumps({'status': 'failed', 'error': f'No quiz with id={qid}'}, default=str),
                            status.HTTP_404_NOT_FOUND
                        )

                except Exception as err:
                    return Response(
                        json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                        status.HTTP_409_CONFLICT
                    )

                filename = f'{qid}.json'
                filepath = os.path.join(tmp_dir, filename)

                if not os.path.exists(filepath):
                    with open(filepath, 'w+') as f:
                        f.write(json.dumps(quiz_data, ensure_ascii=False))

                archive.add(filepath, filename)
            archive.close()

        content = f'/storage/{_dirname}/{archive_name}'
        return Response(content, status.HTTP_200_OK)

    def get(self, request):

        quiz_ids = request.GET.get('ids', None)

        if not quiz_ids:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'ids' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(_) for _ in quiz_ids]

        return self._save_quiz2storage(quiz_ids)


class QuizImportJsonHandlerView(APIView):
    """
    That handler allows to import quizs, exported with QuizExportJsonHandler.
    Or you can put your own 'tar.gz' file with inner quizs json files.
    """

    permission_classes = (IsAuthenticated,)
    # permission_classes = (AllowAny,)
    http_method_names = ['post']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    @staticmethod
    def _process_tar_file(tar_file):

        # wraps bytes to work with it like with file
        file_like_object = tar_file.file

        with tarfile.open(mode='r:gz', fileobj=file_like_object) as tar:
            for quiz in tar.getmembers():
                quiz_data = tar.extractfile(quiz)
                quiz_data = json.loads(quiz_data.read())
                quiz_name = quiz_data.get('name', None)
                questions = quiz_data.get('questions', None)
                if None in [quiz_name, questions]:
                    return Response(
                        json.dumps({'status': 'failed', 'error': "params 'name' and 'questions' is needed"},
                                   default=str),
                        status.HTTP_400_BAD_REQUEST
                    )
                try:
                    DB_CONN.add_quiz(name=quiz_name, questions=questions)
                except Exception as err:
                    return Response(
                        json.dumps({'status': 'failed', 'error': str(err)},
                                   default=str),
                        status.HTTP_409_CONFLICT
                    )

            content = {'status': 'success'}
            return Response(content, status.HTTP_200_OK)

    def post(self, request):

        files = request.FILES

        if not files or not files.get('file'):
            content = {'status': 'no file in payload'}
            return Response(content, status.HTTP_204_NO_CONTENT)

        tar_file = dict(files)['file'][0]

        return self._process_tar_file(tar_file)

class FilledQuizExportHandlerView(APIView):
    """
    This handler allows export filled quiz object into '.xlsx' format file.
    Input param is 'id' which quiz.id in DB.
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.wb = Workbook()

    def create_wb(self, filepath, questions):
        ws = self.wb.active
        ws.title = 'чек-лист'
        cell_range = ws['A1':'F1']
        col_range = ['№', 'Вопрос', 'Ответ', 'Пояснение', 'Ключевой', 'Метка']
        # logger.debug(col_range)
        for c, v in zip(*cell_range, col_range):
            c.value = v

        for i, q in enumerate(questions, 2):
            cell_range = ws[f'A{i}':f'F{i}']
            col_range = [q['sid'], q['text'], q['answer']['value'], q['answer']['description'],
                         q['is_sign'], q['label']]
            # logger.info(col_range)
            for c, v in zip(*cell_range, col_range):
                c.value = v

        self.wb.save(filepath)

    def get(self, request):

        quiz_id = request.GET.get('id', None)

        if not quiz_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_id = int(quiz_id)

        try:
            quiz_data = DB_CONN.get_filled_quiz(quiz_id=quiz_id, current=True)
            quiz_data = quiz_data[0] if quiz_data else None
            if not quiz_data:
                return Response(
                    json.dumps({'status': 'failed', 'error': f'No quiz with id={quiz_id}'}, default=str),
                    status.HTTP_404_NOT_FOUND
                )
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )

        questions = quiz_data['questions']
        quiz_name = quiz_data['name'].replace('«', '"').replace('»', '"')
        quiz_name = quiz_name.replace(' ', '_')
        filled = quiz_data['fill_date'].replace(' ', '').replace(':', '')

        filename = f'{quiz_name}_{filled}.xlsx'
        dirpath = os.path.join(STATIC_CONF['static_path'], 'xlsx')
        filepath = os.path.join(dirpath, filename)

        if os.path.exists(filepath):
            content = f'/xlsx/{filename}'
            return Response(content, status.HTTP_200_OK)

        os.makedirs(dirpath, exist_ok=True)

        self.create_wb(filepath, questions)

        content = f'/xlsx/{filename}'
        return Response(content, status.HTTP_200_OK)
