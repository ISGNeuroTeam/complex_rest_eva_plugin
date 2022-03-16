from rest.permissions import IsAuthenticated
from rest.response import Response, status
from rest.views import APIView
from datetime import datetime
from openpyxl import Workbook
import super_logger
import tempfile
import tarfile
import uuid
import json
import os

from plugins.db_connector.connector_singleton import db
from ..utils.get_user import get_current_user
from ..settings import STATIC_CONF


class QuizsHandlerView(APIView):
    """
    That handler allows to get list of quizs with offset & limit params for pagination.
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._limit = None
        self._offset = None

    @property
    def quizs(self):
        return db.get_quizs(limit=self._limit, offset=self._offset)

    @property
    def quizs_count(self):
        return db.get_quizs_count()

    def get(self, request):

        self._offset, self._limit = request.GET.get('offset', 0), request.GET.get('limit', 10)

        content = {'data': self.quizs, 'count': self.quizs_count}
        return Response(content, status.HTTP_200_OK)


class QuizHandlerView(APIView):
    """
    There is four methods for four actions with quiz objects.
    - get:      returns quiz data by 'id' param;
    - post:     creates new quiz object in DB with data from json body;
    - put:      edit existing quiz object in DB with data from json body by id;
    - delete:   delete existing quiz object from DB by 'id' param;
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get', 'post', 'put', 'delete']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._quiz_id = None

    def get(self, request):

        self._quiz_id = request.GET.get('id', None)

        if not self._quiz_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        try:
            quiz = db.get_quiz(quiz_id=self._quiz_id)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        content = {'data': quiz}
        return Response(content, status.HTTP_200_OK)

    def post(self, request):

        data = request.data
        quiz_name = data.get('name', None)
        questions = data.get('questions', None)

        if None in [quiz_name, questions]:
            return Response(
                json.dumps({'status': 'failed', 'error': "params 'name' and 'questions' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )
        try:
            quiz_id = db.add_quiz(name=quiz_name,
                                  questions=questions)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )
        content = {'data': quiz_id}
        return Response(content, status.HTTP_200_OK)

    def put(self, request):

        data = request.data
        quiz_id = data.get('id', None)
        quiz_name = data.get('name', None)
        questions = data.get('questions', None)

        if not quiz_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        try:
            quiz_id = db.update_quiz(quiz_id=quiz_id,
                                          name=quiz_name,
                                          questions=questions)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )
        content = {'data': quiz_id}
        return Response(content, status.HTTP_200_OK)

    def delete(self, request):

        quiz_id = request.GET.get('id', None)

        if not quiz_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "params 'id' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_id = db.delete_quiz(quiz_id=quiz_id)

        content = {'data': quiz_id}
        return Response(content, status.HTTP_200_OK)


class QuizFilledHandlerView(APIView):
    """
    Handling actions with filled quiz object.
    It's allows two actions:
    - get:      gets filled quiz object from DB by 'id' param and limit/offset params for pagination;
    - post:     adds new fille quiz object to DB with data from json body;
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get', 'post']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def get(self, request):

        quiz_type_id = request.GET.get('id', None)
        offset = request.GET.get('offset', 0)
        limit = request.GET.get('limit', 10)

        try:
            filled_quizs = db.get_filled_quiz(quiz_id=quiz_type_id,
                                              offset=offset,
                                              limit=limit)
            print(filled_quizs)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )

        count = db.get_filled_quizs_count(quiz_type_id) if quiz_type_id else len(filled_quizs)

        content = {'data': filled_quizs, 'count': count}
        return Response(content, status.HTTP_200_OK)

    @staticmethod
    def _iterate_quizs(request, data, filled_ids):
        for quiz_data in data:

            quiz_type_id = quiz_data.get('id', None)
            questions = quiz_data.get('questions', None)

            if None in [quiz_type_id, questions]:
                return Response(
                    json.dumps({'status': 'failed', 'error': "params 'id' and 'questions' is needed"}, default=str),
                    status.HTTP_400_BAD_REQUEST
                )

            try:
                current_user = get_current_user(request)
                if not current_user:
                    return Response(
                        json.dumps({'status': 'failed', 'error': "unauthorized"}, default=str),
                        status.HTTP_401_UNAUTHORIZED
                    )
                db.save_filled_quiz(user_id=None,
                                    user=current_user,
                                    quiz_id=quiz_type_id,
                                    questions=questions)
                filled_ids.append(quiz_type_id)
            except Exception as err:
                return Response(
                    json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                    status.HTTP_409_CONFLICT
                )

        content = {'ids': filled_ids}
        return Response(content, status.HTTP_200_OK)

    def post(self, request):

        filled_ids = list()

        return self._iterate_quizs(request, request.data, filled_ids)


class QuizQuestionsHandlerView(APIView):
    """
    If is need a question list for one or more quiz, this handler for it.
    Input param is 'ids', like '?ids=1,2,3'.
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def get(self, request):

        quiz_ids = request.GET.get('ids', None)

        if not quiz_ids:
            return Response(
                json.dumps({'status': 'failed', 'error': "params 'ids' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(i) for i in quiz_ids if i]

        try:
            questions = db.get_quiz_questions(quiz_ids=quiz_ids)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )

        content = {'data': questions}
        return Response(content, status.HTTP_200_OK)


class QuizExportJsonHandlerView(APIView):
    """
    There is method for export one or more quiz object in '.json' format files.
    Json files returns in 'tar.gz' package with uuid-name.
    """

    permission_classes = (IsAuthenticated,)
    # permission_classes = (AllowAny,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def _save_quiz2storage(self, quiz_ids):

        with tempfile.TemporaryDirectory() as tmp_dir:
            archive_name = f"{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.eva.quiz"
            _dirname = self.handler_id
            _base_path = os.path.join(STATIC_CONF['static_path'], 'storage', _dirname)
            if not os.path.exists(_base_path):
                os.makedirs(_base_path)

            archive_path = os.path.join(_base_path, archive_name)
            archive = tarfile.open(archive_path, mode='x:gz')

            for qid in quiz_ids:
                try:
                    quiz_data = db.get_quiz(quiz_id=qid)
                    if not quiz_data:
                        return Response(
                            json.dumps({'status': 'failed', 'error': f'No quiz with id={qid}'}, default=str),
                            status.HTTP_404_NOT_FOUND
                        )

                except Exception as err:
                    return Response(
                        json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                        status.HTTP_409_CONFLICT
                    )

                filename = f'{qid}.json'
                filepath = os.path.join(tmp_dir, filename)

                if not os.path.exists(filepath):
                    with open(filepath, 'w+') as f:
                        f.write(json.dumps(quiz_data, ensure_ascii=False))

                archive.add(filepath, filename)
            archive.close()

        content = f'/storage/{_dirname}/{archive_name}'
        return Response(content, status.HTTP_200_OK)

    def get(self, request):

        quiz_ids = request.GET.get('ids', None)

        if not quiz_ids:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'ids' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(_) for _ in quiz_ids]

        return self._save_quiz2storage(quiz_ids)


class QuizImportJsonHandlerView(APIView):
    """
    That handler allows to import quizs, exported with QuizExportJsonHandler.
    Or you can put your own 'tar.gz' file with inner quizs json files.
    """

    permission_classes = (IsAuthenticated,)
    # permission_classes = (AllowAny,)
    http_method_names = ['post']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    @staticmethod
    def _process_tar_file(tar_file):

        # wraps bytes to work with it like with file
        file_like_object = tar_file.file

        with tarfile.open(mode='r:gz', fileobj=file_like_object) as tar:
            for quiz in tar.getmembers():
                quiz_data = tar.extractfile(quiz)
                quiz_data = json.loads(quiz_data.read())
                quiz_name = quiz_data.get('name', None)
                questions = quiz_data.get('questions', None)
                if None in [quiz_name, questions]:
                    return Response(
                        json.dumps({'status': 'failed', 'error': "params 'name' and 'questions' is needed"},
                                   default=str),
                        status.HTTP_400_BAD_REQUEST
                    )
                try:
                    db.add_quiz(name=quiz_name, questions=questions)
                except Exception as err:
                    return Response(
                        json.dumps({'status': 'failed', 'error': str(err)},
                                   default=str),
                        status.HTTP_409_CONFLICT
                    )

            content = {'status': 'success'}
            return Response(content, status.HTTP_200_OK)

    def post(self, request):

        files = request.FILES

        if not files or not files.get('file'):
            content = {'status': 'no file in payload'}
            return Response(content, status.HTTP_204_NO_CONTENT)

        tar_file = dict(files)['file'][0]

        return self._process_tar_file(tar_file)


class CatalogsListHandlerView(APIView):

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def get(self, request):

        offset, limit = request.GET.get('offset', 0), request.GET.get('limit', 10)

        catalogs = db.get_catalogs_data(limit=limit, offset=offset)

        content = {'data': catalogs, 'count': db.get_catalogs_count()}
        return Response(content, status.HTTP_200_OK)


class CatalogHandlerView(APIView):

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get', 'post', 'put', 'delete']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def get(self, request):

        catalog_id = request.GET.get('id', None)

        if not catalog_id:
            return Response(
                        json.dumps({'status': 'failed', 'error': "param 'id' is needed"},
                                   default=str),
                        status.HTTP_400_BAD_REQUEST
                    )

        try:
            catalog = db.get_catalog(catalog_id=catalog_id)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)},
                           default=str),
                status.HTTP_400_BAD_REQUEST
            )

        content = {'data': catalog}
        return Response(content, status.HTTP_200_OK)

    def post(self, request):

        data = request.data

        catalog_name = data.get('name', None)
        content = data.get('content', None)

        if None in [catalog_name, content]:
            return Response(
                json.dumps({'status': 'failed', 'error': "params 'name' and 'content' is needed"},
                           default=str),
                status.HTTP_400_BAD_REQUEST
            )

        try:
            catalog_id = db.add_catalog(name=catalog_name,
                                             content=content)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)},
                           default=str),
                status.HTTP_409_CONFLICT
            )

        content = {'id': catalog_id}
        return Response(content, status.HTTP_200_OK)

    def put(self, request):

        data = request.data

        catalog_id = data.get('id', None)
        catalog_name = data.get('name', None)
        content = data.get('content', None)

        if not catalog_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"},
                           default=str),
                status.HTTP_400_BAD_REQUEST
            )

        try:
            catalog_id = db.update_catalog(catalog_id=catalog_id,
                                           name=catalog_name,
                                           content=content)
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)},
                           default=str),
                status.HTTP_409_CONFLICT
            )

        content = {'id': catalog_id}
        return Response(content, status.HTTP_200_OK)

    def delete(self, request):

        catalog_id = request.GET.get('id', None)

        if not catalog_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"},
                           default=str),
                status.HTTP_400_BAD_REQUEST
            )
        catalog_id = db.delete_catalog(catalog_id=catalog_id)

        content = {'id': catalog_id}
        return Response(content, status.HTTP_200_OK)


class FilledQuizExportHandlerView(APIView):
    """
    This handler allows export filled quiz object into '.xlsx' format file.
    Input param is 'id' which quiz.id in DB.
    """

    permission_classes = (IsAuthenticated,)
    http_method_names = ['get']
    handler_id = str(uuid.uuid4())
    logger = super_logger.getLogger('quizs')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.wb = Workbook()

    def create_wb(self, filepath, questions):
        ws = self.wb.active
        ws.title = 'чек-лист'
        cell_range = ws['A1':'F1']
        col_range = ['№', 'Вопрос', 'Ответ', 'Пояснение', 'Ключевой', 'Метка']
        # logger.debug(col_range)
        for c, v in zip(*cell_range, col_range):
            c.value = v

        for i, q in enumerate(questions, 2):
            cell_range = ws[f'A{i}':f'F{i}']
            col_range = [q['sid'], q['text'], q['answer']['value'], q['answer']['description'],
                         q['is_sign'], q['label']]
            # logger.info(col_range)
            for c, v in zip(*cell_range, col_range):
                c.value = v

        self.wb.save(filepath)

    def get(self, request):

        quiz_id = request.GET.get('id', None)

        if not quiz_id:
            return Response(
                json.dumps({'status': 'failed', 'error': "param 'id' is needed"}, default=str),
                status.HTTP_400_BAD_REQUEST
            )

        quiz_id = int(quiz_id)

        try:
            quiz_data = db.get_filled_quiz(quiz_id=quiz_id, current=True)
            quiz_data = quiz_data[0] if quiz_data else None
            if not quiz_data:
                return Response(
                    json.dumps({'status': 'failed', 'error': f'No quiz with id={quiz_id}'}, default=str),
                    status.HTTP_404_NOT_FOUND
                )
        except Exception as err:
            return Response(
                json.dumps({'status': 'failed', 'error': str(err)}, default=str),
                status.HTTP_409_CONFLICT
            )

        questions = quiz_data['questions']
        quiz_name = quiz_data['name'].replace('«', '"').replace('»', '"')
        quiz_name = quiz_name.replace(' ', '_')
        filled = quiz_data['fill_date'].replace(' ', '').replace(':', '')

        filename = f'{quiz_name}_{filled}.xlsx'
        dirpath = os.path.join(STATIC_CONF['static_path'], 'xlsx')
        filepath = os.path.join(dirpath, filename)

        if os.path.exists(filepath):
            content = f'/xlsx/{filename}'
            return Response(content, status.HTTP_200_OK)

        os.makedirs(dirpath, exist_ok=True)

        self.create_wb(filepath, questions)

        content = f'/xlsx/{filename}'
        return Response(content, status.HTTP_200_OK)
