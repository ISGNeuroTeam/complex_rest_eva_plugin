import os
import json
import tarfile
import tempfile
import uuid
import logging

from datetime import datetime
from openpyxl import Workbook

from rest_framework.response import Response
from rest_framework.exceptions import ParseError, APIException, NotFound

from eva_plugin.settings import STATIC_CONF
from eva_plugin.base_handler import BaseHandler
from eva_plugin.quizs.db import db


logger = logging.getLogger('eva_plugin.quizs')


class QuizsHandler(BaseHandler):
    """
    That handler allows to get list of quizs with offset & limit params for pagination.
    """

    def get(self, request):
        _offset = self.get_argument('offset', 0)
        _limit = self.get_argument('limit', 10)

        quizs = db.get_quizs(limit=_limit, offset=_offset)
        return Response({'data': quizs, 'count': db.get_quizs_count()})


class QuizHandler(BaseHandler):
    """
    There is four methods for four actions with quiz objects.
    - get:      returns quiz data by 'id' param;
    - post:     creates new quiz object in DB with data from json body;
    - put:      edit existing quiz object in DB with data from json body;
    - delete:   delete existing quiz object from DB by 'id' param;
    """

    def get(self, request):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise ParseError("param 'id' is needed")
        try:
            quiz = db.get_quiz(quiz_id=quiz_id)
        except Exception as err:
            raise ParseError(str(err))
        return Response({'data': quiz})

    def post(self, request):
        quiz_name = self.data.get('name', None)
        questions = self.data.get('questions', None)
        if None in [quiz_name, questions]:
            raise ParseError("params 'name' and 'questions' is needed")
        try:
            quiz_id = db.add_quiz(name=quiz_name,
                                       questions=questions)
        except Exception as err:
            raise APIException(str(err))
        return Response({'id': quiz_id})

    def put(self, request):
        quiz_id = self.data.get('id', None)
        quiz_name = self.data.get('name', None)
        questions = self.data.get('questions', None)
        if not quiz_id:
            raise ParseError("param 'id' is needed")
        try:
            quiz_id = db.update_quiz(quiz_id=quiz_id,
                                          name=quiz_name,
                                          questions=questions)
        except Exception as err:
            raise APIException(str(err))
        return Response({'id': quiz_id})

    def delete(self, request):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise ParseError("params 'id' is needed")
        quiz_id = db.delete_quiz(quiz_id=quiz_id)
        return Response({'id': quiz_id})


class QuizFilledHandler(BaseHandler):
    """
    Handling actions with filled quiz object.
    It's allows two actions:
    - get:      gets filled quiz object from DB by 'id' param and limit/offset params for pagination;
    - post:     adds new fille quiz object to DB with data from json body;
    """

    def get(self, request):
        quiz_type_id = self.get_argument('id', None)
        offset = self.get_argument('offset', 0)
        limit = self.get_argument('limit', 10)

        try:
            filled_quizs = db.get_filled_quiz(quiz_id=quiz_type_id,
                                                   offset=offset,
                                                   limit=limit)
        except Exception as err:
            raise APIException(str(err))
        count = db.get_filled_quizs_count(quiz_type_id) if quiz_type_id else len(filled_quizs)
        return Response({'data': filled_quizs, 'count': count})

    def post(self, request):
        filled_ids = list()
        for quiz in self.data:
            quiz_type_id = quiz.get('id', None)
            questions = quiz.get('questions', None)

            if None in [quiz_type_id, questions]:
                raise ParseError("params 'id', and 'questions' is needed")
            try:
                db.save_filled_quiz(user_id=self.current_user,
                                         quiz_id=quiz_type_id,
                                         questions=questions)
                filled_ids.append(quiz_type_id)
            except Exception as err:
                raise APIException(str(err))
        return Response({'ids': filled_ids})


class QuizQuestionsHandler(BaseHandler):
    """
    If is need a question list for one or more quiz, this handler for it.
    Input param is 'ids', like '?ids=1,2,3'.
    """

    def get(self, request):
        quiz_ids = self.get_argument('ids', None)
        if not quiz_ids:
            raise ParseError("params 'ids' is needed")
        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(i) for i in quiz_ids if i]
        try:
            logger.debug(quiz_ids)
            questions = db.get_quiz_questions(quiz_ids=quiz_ids)
        except Exception as err:
            raise APIException(str(err))
        return Response({'data': questions})


class FilledQuizExportHandler(BaseHandler):
    """
    This handler allows export filled quiz object into '.xlsx' format file.
    Input param is 'id' which quiz.id in DB.
    """


    def get(self, request):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise ParseError("param 'id' is needed")
        quiz_id = int(quiz_id)
        try:
            quiz_data = db.get_filled_quiz(quiz_id=quiz_id, current=True)
            quiz_data = quiz_data[0] if quiz_data else None
            if not quiz_data:
                raise NotFound(f'No quiz with id={quiz_id}')
        except Exception as err:
            raise APIException(str(err))

        questions = quiz_data['questions']
        quiz_name = quiz_data['name'].replace('«', '"').replace('»', '"')
        quiz_name = quiz_name.replace(' ', '_')
        filled = quiz_data['fill_date'].replace(' ', '').replace(':', '')

        filename = f'{quiz_name}_{filled}.xlsx'
        filepath = os.path.join(STATIC_CONF['static_path'], 'xlsx', filename)
        if os.path.exists(filepath):
            return Response(f'/xlsx/{filename}')

        wb = Workbook()
        ws = wb.active
        ws.title = 'чек-лист'
        cell_range = ws['A1':'F1']
        col_range = ['№', 'Вопрос', 'Ответ', 'Пояснение', 'Ключевой', 'Метка']
        logger.debug(col_range)
        for c, v in zip(*cell_range, col_range):
            c.value = v

        for i, q in enumerate(questions, 2):
            cell_range = ws[f'A{i}':f'F{i}']
            col_range = [q['sid'], q['text'], q['answer']['value'], q['answer']['description'],
                         q['is_sign'], q['label']]
            logger.info(col_range)
            for c, v in zip(*cell_range, col_range):
                c.value = v

        wb.save(filepath)
        return Response(f'/xlsx/{filename}')


class QuizExportJsonHandler(BaseHandler):
    """
    There is method for export one or more quiz object in '.json' format files.
    Json files returns in 'tar.gz' package with uuid-name.
    """

    def get(self, request):
        quiz_ids = self.get_argument('ids', None)
        if not quiz_ids:
            raise ParseError("param 'ids' is needed")
        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(_) for _ in quiz_ids]

        with tempfile.TemporaryDirectory() as tmp_dir:
            archive_name = f"{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.eva.quiz"
            _dirname = str(uuid.uuid4())
            _base_path = os.path.join(STATIC_CONF['static_path'], 'storage', _dirname)
            if not os.path.exists(_base_path):
                os.makedirs(_base_path)

            archive_path = os.path.join(_base_path, archive_name)
            archive = tarfile.open(archive_path, mode='x:gz')

            for qid in quiz_ids:
                try:
                    quiz_data = db.get_quiz(quiz_id=qid)
                    if not quiz_data:
                        raise NotFound(f'No quiz with id={qid}')
                except Exception as err:
                    raise APIException(str(err))

                filename = f'{qid}.json'
                filepath = os.path.join(tmp_dir, filename)

                if not os.path.exists(filepath):
                    with open(filepath, 'w+') as f:
                        f.write(json.dumps(quiz_data, ensure_ascii=False))

                archive.add(filepath, filename)
            archive.close()
        return Response(f'/storage/{_dirname}/{archive_name}')


class QuizImportJsonHandler(BaseHandler):
    """
    That handler allows to import quizs, exported with QuizExportJsonHandler.
    Or you can put your own 'tar.gz' file with inner quizs json files.
    """

    def post(self, request):
        files = request.FILES
        if not files or not files.get('file'):
            return Response({'status': 'no file in payload'})

        tar_file = files['file']

        with tarfile.open(mode='r:gz', fileobj=tar_file) as tar:
            for quiz in tar.getmembers():
                quiz_data = tar.extractfile(quiz)
                quiz_data = json.loads(quiz_data.read())
                quiz_name = quiz_data.get('name', None)
                questions = quiz_data.get('questions', None)
                if None in [quiz_name, questions]:
                    raise ParseError("params 'name' and 'questions' is needed")
                try:
                    db.add_quiz(name=quiz_name,
                                     questions=questions)
                except Exception as err:
                    raise APIException(str(err))
            return Response({'status': 'success'})
